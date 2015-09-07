#!/usr/bin/env python3

# external modules
import pprint
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Style, Alignment
from openpyxl.styles.colors import BLUE

# my modules
import DCC
import get_url_word
import config as cf
import myutil

# Set Excel Styles
# Excel hyperlink style is calibri 11, underline blue
font_url_style = Font(color = BLUE, underline = 'single')
bold_style = Font(bold = True)
align_hv_cen_style = Alignment(horizontal = 'center', vertical = 'center')
align_ver_cen_style = Alignment(vertical = 'center')
align_hv_cen_wrap_style = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)


def cid_ss_headings(ws):
    col = 1
    ws.cell(row = 1, column = col).value = "DOC Handle"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1
    ws.cell(row = 1, column = col).value = "DCC Title"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1
    ws.cell(row = 1, column = col).value = "TMT Number"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1
    ws.cell(row = 1, column = col).value = "Doc Owner"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1    
    ws.cell(row = 1, column = col).value = "CID Ref. Note" 
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style     
    col += 1
    ws.cell(row = 1, column = col).value = "Curr. CID Ref."
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1
    ws.cell(row = 1, column = col).value = "Curr. CID Ref. Comment"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1
    ws.cell(row = 1, column = col).value = "Curr. CID Ref. Owner"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1    
    ws.cell(row = 1, column = col).value = "Curr. CID Ref. Date"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1    
    ws.cell(row = 1, column = col).value = "Sug. CID Ref. (Current)"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style 
    col += 1
    ws.cell(row = 1, column = col).value = "Sug. CID Ref. Comment"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1
    ws.cell(row = 1, column = col).value = "Sug. CID Ref. Owner"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    col += 1    
    ws.cell(row = 1, column = col).value = "Sug. CID Ref. Date"
    ws.cell(row = 1, column = col).alignment = align_hv_cen_wrap_style
    for col in range(1, 14):
        ws.cell(row = 1, column = col).font = bold_style

def cid_report(ss, ss_name):
    # Open the spreadsheet
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    
    # Write and format the headings in Excel
    cid_ss_headings(ws)
    
    ssrow = 2
    for doc in ss:
        col = 1
        # Column 1: DCC Document Handle
        ws.cell(row = ssrow, column = col).value = doc[0]
        ws.cell(row = ssrow, column = col).font = font_url_style
        ws.cell(row = ssrow, column = col).hyperlink = cf.dcc_url + cf.dcc_prop + doc[0]
        ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
        col += 1
        # Column 2: DCC Document Title
        ws.cell(row = ssrow, column = col).value = doc[1]
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        col += 1
        # Column 3: TMT Document Number
        ws.cell(row = ssrow, column = col).value = doc[2]
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
        col += 1
        # Column 4: Document Owner
        ws.cell(row = ssrow, column = col).value = doc[3]
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
        col += 1
        # Column 5: CID Ref. Note
        ws.cell(row = ssrow, column = col).value = doc[4]
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        col += 1
        # Column 6: Current CID Reference Handle
        ws.cell(row = ssrow, column = col).value = doc[5]
        ws.cell(row = ssrow, column = col).font = font_url_style
        ws.cell(row = ssrow, column = col).hyperlink = cf.dcc_url + cf.dcc_prop + doc[5]
        ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
        col += 1
        # Column 7: Current CID Reference Comment
        ws.cell(row = ssrow, column = col).value = doc[6]
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        col += 1
        # Column 8: Current CID Reference Owner
        ws.cell(row = ssrow, column = col).value = doc[7]
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
        col += 1
        # Column 9: Current CID Reference Date
        ws.cell(row = ssrow, column = col).value = datetime.strptime(doc[8], '%a, %d %b %Y %H:%M:%S %Z')
        ws.cell(row = ssrow, column = col).number_format = 'YYYY MMM DD'
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
        col += 1
        # Column 10: Suggested CID Reference Handle
        ws.cell(row = ssrow, column = col).value = doc[9]
        ws.cell(row = ssrow, column = col).font = font_url_style
        ws.cell(row = ssrow, column = col).hyperlink = cf.dcc_url + cf.dcc_prop + doc[9]
        ws.cell(row = ssrow, column = col).alignment = Alignment(vertical = 'center')
        col += 1
        # Column 11: Suggested CID Reference Comment
        ws.cell(row = ssrow, column = col).value = doc[10]
        ws.cell(row = ssrow, column = col).alignment = Alignment(wrap_text = True, vertical = 'center')
        col += 1
        # Column 12: Suggested CID Reference Owner
        ws.cell(row = ssrow, column = col).value = doc[11]
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
        col += 1
        # Column 13: Suggested CID Reference Date
        ws.cell(row = ssrow, column = col).value = datetime.strptime(doc[12], '%a, %d %b %Y %H:%M:%S %Z')
        ws.cell(row = ssrow, column = col).number_format = 'YYYY MMM DD'
        ws.cell(row = ssrow, column = col).alignment = align_hv_cen_style
        col += 1
       
        ssrow += 1
        
    # Set column widths
    
    # Handles:
    ws.column_dimensions["A"].width = 15.0
    ws.column_dimensions["F"].width = 15.0
    ws.column_dimensions["J"].width = 15.0
    # Titles and Comments
    ws.column_dimensions["B"].width = 45.0
    ws.column_dimensions["G"].width = 45.0   
    ws.column_dimensions["K"].width = 45.0 
    # TMT Document Number
    ws.column_dimensions["C"].width = 23.0  
    # Dates
    ws.column_dimensions["I"].width = 20.0    
    ws.column_dimensions["M"].width = 20.0
    # Owner and ref. note.
    ws.column_dimensions["D"].width = 10.0
    ws.column_dimensions["E"].width = 10.0
    ws.column_dimensions["H"].width = 10.0
    ws.column_dimensions["L"].width = 10.0

   # Save the spreadsheet
    wb.save(ss_name)
        
def gen_ss_cid(s, dl):
    # Spreadsheet dataset will be a list of lists
    ss = []
    for d in dl:
        ssrow = []
        print("trying", d)
        if 'Document-' in d:
            # CID link is to a document Handle
            doc = DCC.prop_get(s,d,InfoSet = 'DocAll', WriteProp = True)
            # Now read preferred version
            prefver = DCC.prop_get(s,doc['prefver'],InfoSet = 'VerAll', WriteProp = True)
            # Subject Document
            ssrow.append(doc['dccnum'])
            ssrow.append(doc['dccname'])
            ssrow.append(doc['tmtnum'])
            ssrow.append(doc['owner-username'])  
            # Evaluation of current reference
            ssrow.append('Doc Ref')             
            # Current reference
            ssrow.append(doc['dccnum'])
            ssrow.append(doc['dccname'])
            ssrow.append(doc['owner-username'])   
            ssrow.append(doc['date'])
            # Suggested reference
            ssrow.append(prefver['dccver'])
            ssrow.append(prefver['vercomment'])
            ssrow.append(prefver['owner-username'])
            ssrow.append(prefver['date'])            
            DCC.print_doc_all(doc)
            DCC.print_ver(prefver)
            print(doc['dccnum'], doc['dccname'], doc['tmtnum'], doc['owner-username'])
            print("Doc Ref")
            print(prefver['dccver'], prefver['vercomment'], prefver['owner-username'], prefver['date'])
        elif 'Version-' in d:
            # CID link is to a version handle
            ver = DCC.prop_get(s,d,InfoSet = 'VerAll', WriteProp = True)
            doc = DCC.prop_get(s,ver['dccdoc'],InfoSet = 'DocAll', WriteProp = True)
            # Subject Document
            ssrow.append(doc['dccnum'])
            ssrow.append(doc['dccname'])
            ssrow.append(doc['tmtnum'])
            ssrow.append(doc['owner-username'])  
            # find info on the preferred version
            prefver = DCC.prop_get(s,doc['prefver'],InfoSet = 'VerAll', WriteProp = True)
            # Evaluation of current reference          
            print("CID references Version")
            print(doc['dccnum'], doc['dccname'], doc['tmtnum'])
            print("Referenced Version")
            print(ver['dccver'], ver['vercomment'], ver['owner-username'], ver['date'])
            if prefver['dccver'] == ver['dccver']:
                print("Referenced Version IS the Preferred Version")
                # Evaluation of current reference
                ssrow.append('Pref. Ver.') 
                DCC.print_doc_all(doc)
                DCC.print_ver(prefver)
                # Current reference
                ssrow.append(prefver['dccver'])
                ssrow.append(prefver['vercomment'])
                ssrow.append(prefver['owner-username'])
                ssrow.append(prefver['date'])  
            else:
                print("Referenced Version IS NOT the Preferred Version") 
                # Evaluation of current reference
                ssrow.append('Non-Pref. Ver.') 
                print("Non-Preferred Version")
                print(prefver['dccver'], prefver['vercomment'], prefver['owner-username'], prefver['date'])    
                DCC.print_doc_all(doc)
                DCC.print_ver(ver)  
                DCC.print_ver(prefver) 
                # Current reference
                ssrow.append(ver['dccver'])
                ssrow.append(ver['vercomment'])
                ssrow.append(ver['owner-username'])
                ssrow.append(ver['date']) 
            # Suggested reference
            ssrow.append(prefver['dccver'])
            ssrow.append(prefver['vercomment'])
            ssrow.append(prefver['owner-username'])
            ssrow.append(prefver['date'])           
        else:
            print("Don't know how to handle", d)
            sys.exit()  
        ss.append(ssrow)
    return(ss)

def get_cid_ssdata(s, json_handlelist, json_ssdata):
    fh = open(json_handlelist,'r')
    dl = json.load(fh)
    fh.close()
    ss = gen_ss_cid(s, dl)
    fh = open(json_ssdata,'w')
    json.dump(ss,fh)
    fh.close()

def write_spreadsheet(json_fname, xls_fname):
    # Open the spreadsheet
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    fh = open(json_fname,'r')
    ss = json.load(fh)
    fh.close()
    cid_report(ss,xls_fname)

def make_handle_lists(json_ssdata, docfname, verfname):
    fh = open(json_ssdata,'r')
    ss = json.load(fh)
    fh.close()
    dl = []
    vl = []
    for doc in ss:
        dl.append(doc[0])
        vl.append(doc[9])
    dfh = open(docfname,'w')
    vfh = open(verfname,'w')
    json.dump(dl,dfh)
    json.dump(vl,vfh)
    dfh.close()
    vfh.close()
    
def make_cid(dirpath, CID_coll, htmlfile, outroot):
    # Get list of DCC documents from a Word file saved as html
    get_url_word.get_url_word(dirpath + outroot, dirpath + htmlfile)  

    # Login to DCC
    s = DCC.login(cf.dcc_url + cf.dcc_login)

    json_handlelist = dirpath + outroot + 'bothlist.txt'
    json_ssdata = dirpath + outroot + 'CID.txt'
    get_cid_ssdata(s, json_handlelist, json_ssdata)

    xlfile = dirpath + outroot + 'CID_Analysis.xls'
    write_spreadsheet(json_ssdata, xlfile)

    json_verlist = dirpath + outroot + 'ver_list.txt'
    json_doclist = dirpath + outroot + 'doc_list.txt'
    make_handle_lists(json_ssdata, json_doclist, json_verlist)

    ## Remove the files that are currently located in the collection for the CID

    if myutil.get_yn('Remove location (not delete) of files from ' + CID_coll[0] +'(Y/N)?: '):
        doclist = DCC.list_obj_in_coll(s, CID_coll[0],Print=True,Jwrite=False,Depth='infinity',Type='Doc',WriteProp=False)
        for doc in doclist:
            DCC.dcc_remove_doc_from_coll(s, doc, CID_coll[0])

    ## Add CID files to the collection
    fh = open(json_doclist, 'r')
    dl = json.load(fh)
    fh.close()

    if myutil.get_yn('Add CID files to ' + CID_coll[0] +' (Y/N)?: '):
        DCC.add_docs_2_collections(s, dl, CID_coll)

    # Check that the expected docs are in the collection
    DCC.check_docs_in_coll(s, dl, CID_coll)